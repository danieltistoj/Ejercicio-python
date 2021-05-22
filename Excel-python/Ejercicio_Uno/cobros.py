import openpyxl, pprint

libro = openpyxl.load_workbook(filename='cobros.xlsx')
hoja = libro['Hoja1']

clientes = {}

for fila in range(2,hoja.max_row+1):

    cliente = hoja.cell(row=fila,column=1).value
    monto = hoja.cell(row=fila,column=4).value

    clientes.setdefault(cliente, 0) #esto es si el cliente no existe
    clientes[cliente]+=monto  #si existe se le suma el monto

#pprint.pprint(clientes)

for fila in range(2,hoja.max_row+1):

    cliente = hoja.cell(row=fila,column=1).value
    total = hoja.cell(row=fila,column=5)

    total.value = clientes[cliente]

libro.save('cobrar_v2.xlsx')


