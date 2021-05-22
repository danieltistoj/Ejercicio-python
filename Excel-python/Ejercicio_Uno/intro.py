from openpyxl import Workbook
#podemos crear un archivo de excel

libro = Workbook()
#un libro debe de tener minimo una hoja
hoja1= libro.active

#print(hoja1)

#Podemos crear una hoja nueva
hoja2 = libro.create_sheet("hoja2")

#print(hoja1.title)
#cambiar nombre
hoja1.title = 'hoja1'

#print(hoja1.title)

#print(libro['hoja1'].title)

#acceder a todas las hojas
#print(libro.sheetnames)

#acceder a cada una de las hojas

#for hoja in libro:
#    print(hoja.title)

hoja2_copia = libro.copy_worksheet(hoja2)

#print(libro.sheetnames)

#acceder a una celda

celda_a1_hoja1 = hoja1['A1']
celda_a1_hoja1.value = "Suscribete al canal"
#print(celda_a1_hoja1.value)
#otra forma  de mostrar celdas
#print(hoja1.cell(row=1,column=1).value)

#Acceder a una fila completa
#por rango
rango = hoja1['A1':'B2']
#print(rango)
#Acceder columna
rango = hoja1['A':'A']
#print(rango)
#Cuando una celda es creada no todas las celdas son creadas

#acceder a filas y columnas en su totalidad

#... Acceder a filas

for fila in hoja1.iter_rows(min_row=1,max_col=3,max_row=2):
    print(fila)

#Acceder a una fila en cada iteracion de este loop
for fila in hoja1.iter_rows(min_row=1,max_col=3,max_row=2):
    #Acceder a una celda de la fila en la que estamos en cada iteracion de esta loop
    for celda in fila:
        #Si queremos ver el valor que cada una de las celdas solo debemos de poner celda.value
        print(celda)

#Para las columnas
print("COLUMNAS")
#Con values_only pasa los valores de cada celda
for columna in hoja1.iter_cols(min_row=1,max_col=3,max_row=2,values_only=True):
    print(columna)

for fila in hoja1.iter_cols(min_row=1,max_col=3,max_row=2):
    #Acceder a una celda de la fila en la que estamos en cada iteracion de esta loop
    for celda in fila:
        #Si queremos ver el valor que cada una de las celdas solo debemos de poner celda.value
        print(celda)

#libro.save('.\prueba.xlsx')
#ver maximos y minimos

#print(hoja1.max_row,hoja1.max_column,hoja1.min_row,hoja1.min_column)
#from openpyxl import load_workbook
#libro_nuevo = load_workbook('cobros.xlsx')
#print(libro_nuevo.sheetnames)