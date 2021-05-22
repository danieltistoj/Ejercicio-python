from openpyxl import Workbook
from openpyxl.styles import Font #Nos permite cambiar el formato
import time

libro = Workbook()
hoja = libro.active # indicamos la hoja activa

hoja['A1'] = 5
hoja['A2'] = 10

hoja['B1'] = 'rango'
hoja['B1'].font = Font(color='FF0000',bold=True)

for i in range(2,15):
    hoja[f'B{i}'] = i**2

hoja2 = libro.create_sheet('hoja2')
hoja2['A1'] = 'Hola mundo'
fecha = time.strftime('%x')
hoja2['A2']=fecha

hoja3 = libro.create_sheet('hoja3')
hoja3.merge_cells('A1:D1')
#Quitar union
#hoja3.unmerge_cells('A1:D1')
hoja3['A1'] = 'Prueba de union de celdas'
libro.save('prueba.xlsx')
